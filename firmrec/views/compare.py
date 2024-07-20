import os

import openpyxl
from openpyxl.utils import get_column_letter


COL_VULN = 1
COL_VENDOR = 2
COL_FIRMWARE = 3
COL_PATH = 4
COL_ENTRY_ADDR = 5
COL_VULN_ADDR = 6
COL_FIRMREC = 7
COL_GENIUS = 8
COL_SCORE = 9
COL_RANK = 10

COLS = [
    COL_VULN,
    COL_VENDOR,
    COL_FIRMWARE,
    COL_PATH,
    COL_ENTRY_ADDR,
    COL_VULN_ADDR,
    COL_FIRMREC,
    COL_GENIUS,
    COL_SCORE,
    COL_RANK,
]
TITLES = [
    "漏洞",
    "厂家",
    "Firmware",
    "路径",
    "Entry地址",
    "Vuln地址",
    "FirmRec",
    "Genius Top 100",
    "Genius Score",
    "Genius Rank",
]


class ComparisonXlsxView:
    def __init__(self, cmp):
        self.cmp = cmp

    def save_xlsx(self, output_dir):
        wb = openpyxl.Workbook()
        sub_sheet = wb.active
        sub_sheet.title = "Compare"

        for col, title in zip(COLS, TITLES):
            self.set_cell(sub_sheet, 1, col, title)

        cur_line = 2
        for vuln_name, vuln_reports in self.cmp.items():
            self.set_cell(sub_sheet, cur_line, COL_VULN, vuln_name)
            score_ranks = dict()
            score_count = 0
            for k, vuln_report in sorted(
                vuln_reports.items(),
                key=lambda x: x[1]["genius_score"]
                if x[1]["genius_score"] is not None
                else -1.0,
            ):
                self.set_cell(sub_sheet, cur_line, COL_VENDOR, k.vendor)
                self.set_cell(sub_sheet, cur_line, COL_FIRMWARE, k.firmware_id)
                self.set_cell(sub_sheet, cur_line, COL_PATH, k.path)
                entry_addr = vuln_report["entry_addr"]
                entry_addr = (
                    entry_addr
                    if entry_addr is None or isinstance(entry_addr, str)
                    else hex(entry_addr)
                )
                vuln_addr = vuln_report["vuln_addr"]
                vuln_addr = (
                    vuln_addr
                    if vuln_addr is None or isinstance(vuln_addr, str)
                    else hex(vuln_addr)
                )
                self.set_cell(sub_sheet, cur_line, COL_ENTRY_ADDR, entry_addr)
                self.set_cell(sub_sheet, cur_line, COL_VULN_ADDR, vuln_addr)
                self.set_cell(sub_sheet, cur_line, COL_FIRMREC, vuln_report["firmrec"])
                self.set_cell(sub_sheet, cur_line, COL_GENIUS, vuln_report["genius"])
                if vuln_report["genius_score"] is not None:
                    score_count += 1
                    score = vuln_report["genius_score"]
                    if score in score_ranks:
                        rank = score_ranks[score]
                    else:
                        rank = score_ranks[score] = score_count
                    self.set_cell(sub_sheet, cur_line, COL_SCORE, score)
                    self.set_cell(sub_sheet, cur_line, COL_RANK, rank)
                cur_line += 1

        output_path = os.path.join(output_dir, "compare.xlsx")
        wb.save(output_path)

    @classmethod
    def set_cell(
        cls,
        sheet,
        row,
        col,
        value,
        fontcolor="00000000",
        fillcolor=None,
        bold=False,
        leftborder=False,
        rightborder=False,
    ):
        sheet.cell(row=row, column=col).value = value
        sheet.cell(row=row, column=col).font = openpyxl.styles.Font(
            "Consolas", color=fontcolor, bold=bold
        )
        if fillcolor:
            sheet.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                "solid", fgColor=fillcolor
            )



class CSVResultSet:
    def __init__(self) -> None:
        pass


class ComparisonCSVView:
    def __init__(self, cmp):
        self.cmp = cmp
