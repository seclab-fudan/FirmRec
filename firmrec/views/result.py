"""
View of vulnerable result
"""
from __future__ import annotations
from typing import TYPE_CHECKING, Any
from functools import cached_property

import openpyxl
from openpyxl.utils import get_column_letter
import csv

from .firmdb import FirmwareStatistic
from ..models.result import ResultKey
from ..models.target_info import TargetInfo
from .vuln_result import VulnResultItemView, VulnFilter

if TYPE_CHECKING:
    from ..models.result import ResultSet


__all__ = ["ResultSetXlsxView"]

RED = "00FF0000"
DARKRED = "00C00000"
BLACK = "00000000"
WHITE = "00FFFFFF"
GREEN = "0000FF00"
DARKGREEN = "0000AA00"
BLUE = "000000FF"
DARKBLUE = "00336699"
YELLOW = "00EEEA96"
ORANGE = "00e69138"
GREY = "00d9d9d9"

COL_VENDOR = 1
COL_FIRMWARE = 2
COL_PRODUCT = 3
COL_VERSION = 4
COL_PATH = 5
COL_VULN_NAME = 6
COL_ENTRY_ADDR = 7
COL_TIME = 8
COL_TIMEOUT = 9
COL_GT = 10
COL_RESULT = 11
COL_RESERVE = 12
COL_VERIFY = 13
COL_FUNC_NAME = 14
COL_NEW_VERSION = 15
COL_DISCLOSE = 16
COL_ONSELL = 17
COL_REPORT = 18
COL_REPORT_URL = 19
COL_DESCRIPTION = 20
COL_KNOWN_VULN = 21
COL_SOURCE = 22
COL_KEYWORD = 23
COL_SINK = 24
COL_REASON = 25
COL_REPEAT = COL_FINAL = 26

COLS = [
    COL_VENDOR,
    COL_FIRMWARE,
    COL_PRODUCT,
    COL_VERSION,
    COL_PATH,
    COL_VULN_NAME,
    COL_ENTRY_ADDR,
    COL_TIME,
    COL_TIMEOUT,
    COL_GT,
    COL_RESULT,
    COL_VERIFY,
    COL_RESERVE,
    COL_FUNC_NAME,
    COL_NEW_VERSION,
    COL_DISCLOSE,
    COL_ONSELL,
    COL_REPORT,
    COL_REPORT_URL,
    COL_DESCRIPTION,
    COL_KNOWN_VULN,
    COL_SOURCE,
    COL_KEYWORD,
    COL_SINK,
    COL_REASON,
]
ATTRS = [
    "vendor",
    "firmware_id",
    None,
    None,
    "path",
    "vuln_name",
    "entry_addr",
    "run_time",
    "timeout",
    None,
    None,
]
TITLES = [
    "厂家",
    "Firmware",
    "产品",
    "版本",
    "路径",
    "漏洞",
    "入口地址",
    "时间",
    "超时",
    "参考结果",
    "结果",
    "",
    "Verify",
    "函数名",
    "固件最新版本",
    "官方披露",
    "官方在售",
    "漏洞披露情况",
    "漏洞报告链接",
    "备注",
    "可能已知漏洞",
    "Source",
    "Keyword",
    "Sink",
    "漏洞原因",
]


class ResultSetXlsxView:
    """
    Xlsx table view of FirmRec result set
    """

    def __init__(self, resset: ResultSet, migrate_from=None) -> None:
        self.resset = resset
        self.wb = self.get_workbook(resset, migrate_from)

    def save(self, output_path):
        """Save workbook to output_path"""
        self.wb.save(output_path)

    @classmethod
    def get_migrate_data(cls, wb_path):
        """
        Get migrate data from xlsx file
        """
        if ":" in wb_path:
            wb_path, sheet_name = wb_path.split(":", 1)
        else:
            sheet_name = "Result"
        m_wb = openpyxl.load_workbook(wb_path)
        sheet = m_wb[sheet_name]

        # estimate max rows
        migrate_data = {}  # migrate data
        rows = sheet.rows
        next(rows)  # skip title
        curr_row_idx = 2
        min_row = 2
        # max_row = cls._infer_max_row(sheet)
        max_row = sheet.max_row
        max_col = COL_FINAL
        for row in sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=1, max_col=max_col
        ):
            if not row[0].value:
                break
            refer_id = cls._get_refer_id_from_row(sheet, curr_row_idx)
            if refer_id not in migrate_data:
                migrate_data[refer_id] = []
            md_row = [
                cls.get_cell(sheet, row, col)
                for col in range(COL_VENDOR, COL_FINAL + 1)
            ]
            migrate_data[refer_id].append(md_row)
            curr_row_idx += 1

        m_wb.close()
        return migrate_data

    @classmethod
    def _infer_max_row(cls, sheet):
        min_row = max_row = 2
        while True:
            if not sheet.cell(max_row, 1).value:
                break
            max_row += 1
        return min_row

    @classmethod
    def _get_refer_id_from_row(cls, sub_sheet, row_idx):
        cols = [COL_VENDOR, COL_FIRMWARE, COL_PATH, COL_VULN_NAME, COL_ENTRY_ADDR]
        attrs = ["vendor", "firmware_id", "path", "vuln_name", "entry_addr"]
        key_d = {
            attr: cls.get_cell(sub_sheet, row_idx, col)
            for col, attr in zip(cols, attrs)
        }
        key = ResultKey.from_dict(key_d)
        refer_id = TargetInfo.refer_id(key)
        return refer_id

    @classmethod
    def _sort_keys(cls, keys: list[ResultKey]):
        return sorted(
            keys,
            key=lambda key: (
                key.vendor,
                key.firmware_id,
                key.path,
                key.entry_addr,
                key.vuln_name,
                key.extra or 0,
            ),
        )

    @classmethod
    def _sort_results(cls, resset: ResultSet):
        return [resset[key] for key in cls._sort_keys(resset.keys())]

    @classmethod
    def get_workbook(cls, resset: ResultSet, migrate_from=None):
        """
        Generate a workbook from resset

        :param resset: result set
        :param migrate_from: the path of xlsx file that generated previously
        :return: generated workbook
        """
        wb = openpyxl.Workbook()
        sub_sheet = wb.active
        sub_sheet.title = "Result"

        firmst = FirmwareStatistic()

        if migrate_from:
            migrate_data = cls.get_migrate_data(migrate_from)

        sub_sheet.column_dimensions[get_column_letter(COL_VULN_NAME)].width = 15
        for col, title in zip(COLS, TITLES):
            cls.set_cell(sub_sheet, 1, col, title)

        cur_line = 2
        for res in cls._sort_results(resset):
            cur_row = [None for _ in range(COL_FINAL + 1)]
            firmware_identifier = f"{res.vendor}/{res.firmware_id}"
            db_infos = list(firmst.query({"image.filename": firmware_identifier}))

            if db_infos:
                cls.set_cell(
                    sub_sheet,
                    cur_line,
                    COL_PRODUCT,
                    db_infos[0]["product"],
                    cur_row=cur_row,
                )
                cls.set_cell(
                    sub_sheet,
                    cur_line,
                    COL_VERSION,
                    db_infos[0]["version"],
                    cur_row=cur_row,
                )

            for col, attr in zip(COLS, ATTRS):
                if not attr:
                    continue
                val = getattr(res, attr)
                if col == COL_ENTRY_ADDR:
                    val = hex(val)
                elif col == COL_TIME:
                    val = str(round(val)) + "s"
                else:
                    val = str(val)
                cls.set_cell(sub_sheet, cur_line, col, val, cur_row=cur_row)
            cls.set_cell(
                sub_sheet,
                cur_line,
                COL_RESULT,
                "VULN" if res.vuln else "NORM",
                cur_row=cur_row,
            )

            # extract vuln_detail
            if res.vuln:
                vuln_view = VulnResultItemView(res)
                cls.set_cell(
                    sub_sheet, cur_line, COL_SOURCE, vuln_view.source, cur_row=cur_row
                )
                cls.set_cell(
                    sub_sheet, cur_line, COL_KEYWORD, vuln_view.keyword, cur_row=cur_row
                )
                cls.set_cell(
                    sub_sheet, cur_line, COL_SINK, vuln_view.sink, cur_row=cur_row
                )
                cls.set_cell(
                    sub_sheet,
                    cur_line,
                    COL_REASON,
                    vuln_view.vuln_reason,
                    cur_row=cur_row,
                )

            if migrate_from:
                ref_id = TargetInfo.refer_id(res.key)
                md_rows = migrate_data.get(ref_id, [])
                if not md_rows:
                    cur_line += 1
                    continue
                max_count = 0
                max_count_md_row = None
                for md_row in md_rows:
                    count = 0
                    for idx, md_data in enumerate(md_row):
                        col = idx + 1
                        data = cls.get_cell(sub_sheet, cur_row, col)
                        if md_data and data and md_data == data:
                            count += 1
                    if count > max_count:
                        max_count = count
                        max_count_md_row = md_row
                for idx, data in enumerate(max_count_md_row):
                    col = idx + 1
                    if not data:
                        continue
                    cls.set_cell(sub_sheet, cur_line, col, data)

            cur_line += 1
        return wb

    @classmethod
    def get_cell(cls, sheet, row, col):
        """Get cell in sheet"""
        if isinstance(row, int):
            val = sheet.cell(row=row, column=col).value
        else:
            val = row[col - 1]
            if hasattr(val, "value"):
                val = val.value
        if val:
            if col == COL_FIRMWARE:
                if isinstance(val, float):
                    val = round(val)
                if isinstance(val, int):
                    val = str(val)
        return val

    @classmethod
    def set_cell(
        cls,
        sheet,
        row,
        col,
        value,
        fontcolor="00000000",
        fillcolor=None,
        bold=False,
        leftborder=False,
        rightborder=False,
        cur_row=None,
    ):
        """Set cell in sheet"""
        sheet.cell(row=row, column=col).value = value
        sheet.cell(row=row, column=col).font = openpyxl.styles.Font(
            "Consolas", color=fontcolor, bold=bold
        )
        if fillcolor:
            sheet.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                "solid", fgColor=fillcolor
            )
        if leftborder:
            sheet.cell(row=row, column=col).border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style="thick", color=RED)
            )
        if rightborder:
            sheet.cell(row=row, column=col).border = openpyxl.styles.Border(
                right=openpyxl.styles.Side(border_style="thick", color=RED)
            )
        if cur_row:
            cur_row[col - 1] = value


class CSVResultItem:
    """Result item of FirmRec result set"""

    def __init__(self, row) -> None:
        self.row = tuple(row)
        self.gt = None

    def set_gt(self, gt):
        """Set ground truth VULN/NORM"""
        assert gt in ("VULN", "NORM")
        self.gt = gt

    @property
    def key(self):
        """Unique key of result item"""
        return (self.vendor, self.firmware_id, self.path, self.entry_addr, self.keyword)

    @property
    def firmware_id(self):
        """firmware_id"""
        return self.row[0]

    @property
    def vendor(self):
        """vendor"""
        return self.row[1]

    @property
    def path(self):
        """bin_path"""
        return self.row[2]

    @property
    def entry_addr(self):
        """entry_addr"""
        return self.row[3] if isinstance(self.row[3], int) else int(self.row[3])

    @property
    def source(self):
        """source"""
        return self.row[4]

    @property
    def keyword(self):
        """ "keyword"""
        return self.row[5]

    @property
    def sink(self):
        """sink"""
        return self.row[6]

    @property
    def vuln_reason(self):
        """vuln_reason"""
        return self.row[7]

    @property
    def run_time(self):
        """run_time"""
        return float(self.row[8])

    @property
    def result(self):
        """Result VULN/NORM"""
        return self.row[-1]

    @property
    def vuln(self):
        """Is vulnerable"""
        return self.result == "VULN"

    @property
    def correct(self):
        """Is correct"""
        assert self.gt is not None
        return self.gt == self.result

    def __eq__(self, __value: object) -> bool:
        if not isinstance(__value, CSVResultItem):
            return False
        return self.key == __value.key

    def __hash__(self) -> int:
        return hash(self.key)


class CSVResultSet:
    """CSV result set of FirmRec result set"""

    def __init__(self):
        self._results = {}
        self._indexed_result_sets = {}

    def add(self, item: CSVResultItem):
        """Add result item"""
        self._results[item.key] = item
    
    def pop(self, item: CSVResultItem):
        """Pop result item"""
        if item.key not in self._results:
            return None
        return self._results.pop(item.key)

    def get(self, key, default=None):
        """Get result item"""
        return self._results.get(key, default)

    def keys(self):
        """Keys of result set"""
        return self._results.keys()
    
    def values(self):
        """Values of result set"""
        return self._results.values()

    def find_results(self, **kwargs):
        """Find results in result set with give field filter.

        :param kwargs: key and value to match
            :field filter: function to filter result
        :return: a generator of found results that matching field filters,
            if no filter is give, all results will be returned
        """
        filter = kwargs.pop("filter", None)
        indexed_result_sets = []
        for key, value in kwargs.items():
            if not hasattr(CSVResultItem, key):
                raise ValueError(f"Invalid key {key}")
            
            # build index
            if key not in self._indexed_result_sets:
                self._indexed_result_sets[key] = val_index = {}
                for r in self._results.values():
                    r_value = getattr(r, key)
                    if r_value not in val_index:
                        val_index[r_value] = set()
                    val_index[r_value].add(r)
            if value in self._indexed_result_sets[key]:
                result_sets = self._indexed_result_sets[key][value]
                indexed_result_sets.append(result_sets)
        if indexed_result_sets:
            results = set.intersection(*indexed_result_sets)
        else:
            results = self.values()

        for result in results:
            for key, value in kwargs.items():
                if getattr(result, key) != value:
                    break
            else:
                if filter and not filter(result):
                    continue
                yield result

    def find_resset(self, **kwargs) -> CSVResultSet:
        """
        Find results in result set with give field filter.
        :param kwargs: key and value to match
        :return: a new result set with found results that matching field filters,
        """
        resset = CSVResultSet()
        for result in self.find_results(**kwargs):
            resset.add(result)
        return resset

    def __iter__(self):
        return iter(self._results.values())

    def __len__(self):
        return len(self._results)

    def __contains__(self, key):
        if isinstance(key, CSVResultItem):
            key = key.key
        else:
            try:
                key = CSVResultItem(key)
                key = key.key
            except ValueError:
                return False
            except TypeError:
                return False
        return key in self._results

    def __getitem__(self, key):
        if isinstance(key, CSVResultItem):
            key = key.key
        elif isinstance(key, tuple):
            key = key[:6]
        else:
            return None
        return self._results.get(key, None)

    def __setitem__(self, key, value):
        if isinstance(key, CSVResultItem):
            key = key.key
        elif isinstance(key, tuple):
            key = key[:6]
        else:
            return None
        self._results[key] = value

    def __sub__(self, another: CSVResultSet):
        """Subtract another result set"""
        if not isinstance(another, CSVResultSet):
            raise ValueError(f"Can't sub {type(another)} from {type(self)}")
        resset = CSVResultSet()
        for r in self:
            if r not in another:
                resset.add(r)
        return resset

    @cached_property
    def T(self):
        """Vulnerable results"""
        resset = CSVResultSet()
        for r in self:
            if r.result == "VULN":
                resset.add(r)
        return resset

    @cached_property
    def F(self):
        """Normal results"""
        resset = CSVResultSet()
        for r in self:
            if r.result == "NORM":
                resset.add(r)
        return resset

    @cached_property
    def TP(self):
        """True positive results"""
        resset = CSVResultSet()
        for r in self:
            if r.gt and r.result == "VULN" and r.correct:
                resset.add(r)
        return resset

    @cached_property
    def FP(self):
        """False positive results"""
        resset = CSVResultSet()
        for r in self:
            if r.gt and r.result == "VULN" and not r.correct:
                resset.add(r)
        return resset

    @cached_property
    def TN(self):
        """True negative results"""
        resset = CSVResultSet()
        for r in self:
            if r.gt and r.result == "NORM" and r.correct:
                resset.add(r)
        return resset

    @cached_property
    def FN(self):
        """False negative results"""
        resset = CSVResultSet()
        for r in self:
            if r.gt and r.result == "NORM" and not r.correct:
                resset.add(r)
        return resset

    def union(self, another: CSVResultSet):
        """Union with another result set"""
        if not isinstance(another, CSVResultSet):
            raise ValueError(f"Can't union {type(another)} with {type(self)}")
        resset = CSVResultSet()
        for r in self:
            resset.add(r)
        for r in another:
            if r.gt and not r.correct:
                continue
            resset.add(r)
        return resset

    def diff(self, another: CSVResultSet):
        """Diff with another result set

        :param another: resultset to diff
        :param diff_func: function to distinguish different result
        :return: (resset_l, resset_r, resset_vl, resset_vr)
        """
        if not isinstance(another, CSVResultSet):
            raise ValueError(f"Can't sub {type(another)} from {type(self)}")
        resset_l = CSVResultSet()
        resset_r = CSVResultSet()
        resset_vl = CSVResultSet()
        resset_vr = CSVResultSet()
        keys = set(self.keys()).union(another.keys())
        for k in keys:
            r = self.get(k, None)
            r_c = another.get(k, None)
            if not r:
                resset_r.add(r_c)
            elif not r_c:
                resset_l.add(r)
            else:
                if r.vuln and not r_c.vuln:
                    resset_vl.add(r)
                elif not r.vuln and r_c.vuln:
                    resset_vr.add(r_c)

        return resset_l, resset_r, resset_vl, resset_vr

    def __str__(self) -> str:
        return f"CSVResultSet(A={len(self)}, T={len(self.T)}, F={len(self.F)}, TP={len(self.TP)}, TN={len(self.TN)}, FP={len(self.FP)}, FN={len(self.FN)})"

    def __repr__(self) -> str:
        return str(self)

    @classmethod
    def load(cls, path, gt: CSVResultSet = None):
        """Load result set from csv file"""
        resset = cls()
        resset_dup = cls()
        if not gt:
            gt = cls()
        with open(path, "r", encoding="utf-8") as fp:
            reader = csv.reader(fp, delimiter=",")
            for row in reader:
                if not row:
                    continue
                item = CSVResultItem(row)
                item_gt = gt.get(item.key, None)
                if item_gt:
                    item.set_gt(item_gt.result)
                if item not in resset:
                    resset.add(item)
                elif item.vuln:
                    item_dup = resset.pop(item)
                    resset_dup.add(item_dup)
                    resset.add(item)
                else:
                    resset_dup.add(item)
        return resset, resset_dup


class ResultSetCSVView:
    """CSV table view of FirmRec result set"""

    def __init__(self, resset: ResultSet) -> None:
        self.resset = resset

    def save(self, output_path):
        """Save csv to output_path"""
        vf = VulnFilter()
        output_fp = open(output_path, "w+", encoding="utf-8")
        writer = csv.writer(output_fp, delimiter=",", lineterminator="\n")

        title = [
            "firmware_id",
            "vendor",
            "bin_path",
            "entry_addr",
            "source",
            "keyword",
            "sink",
            "vuln_reason",
            "run_time",
            "vuln",
        ]
        writer.writerow(title)

        visited = set()
        for r in self.resset.results:
            if r.vuln:
                r = r.load()
            rid = TargetInfo.refer_id(r)
            # if rid in visited:
            #     continue
            # visited.add(rid)
            line = list(rid)
            line.pop()

            if vf.is_vulnerable(r):
                vv = VulnResultItemView(r)
                line.append(vv.sink)
                line.append(vv.vuln_reason)
                line.append(r.run_time)
                line.append("VULN")
            else:
                line.append("")
                line.append("")
                line.append(r.run_time)
                line.append("NORM")

            writer.writerow(line)
